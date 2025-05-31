import os
import requests
import openpyxl
from urllib.parse import urlparse
from pathlib import Path

# === CONFIGURATION ===
EXCEL_FILE = r"D:\Projects\Cyber Projects\Image Download Script\products_list_for_uber_v2.xlsx"
COLUMN_WITH_URLS = 'G'              # Excel column where URLs are stored
START_ROW = 2                       # Assuming row 1 is a header
DOWNLOAD_DIR = "downloaded_images"  # Folder to save images

# === PREPARE FOLDER ===
Path(DOWNLOAD_DIR).mkdir(parents=True, exist_ok=True)

# === LOAD EXCEL ===
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active

# === DOWNLOAD LOOP ===
for row in range(START_ROW, ws.max_row + 1):
    cell = f"{COLUMN_WITH_URLS}{row}"
    url = ws[cell].value

    if not url:
        continue

    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()

        # Extract file name from URL or generate one
        parsed_url = urlparse(url)
        filename = os.path.basename(parsed_url.path)
        if not filename or '.' not in filename:
            filename = f"image_{row}.jpg"

        filepath = os.path.join(DOWNLOAD_DIR, filename)

        with open(filepath, 'wb') as f:
            f.write(response.content)

        print(f"Downloaded: {filename}")
    except Exception as e:
        print(f"Failed to download from {url} (row {row}): {e}")
